from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Loan Calculator"

# ======================
# INPUT SECTION
# ======================

ws["A1"] = "Home Loan Amount"
ws["B1"] = 1000000

ws["A2"] = "Interest Rate (%)"
ws["B2"] = 10

ws["A3"] = "Loan Tenure (Months)"
ws["B3"] = 120

for cell in ["A1","A2","A3"]:
    ws[cell].font = Font(bold=True)

# ======================
# DASHBOARD
# ======================

ws["D1"] = "Scheduled Monthly EMI"
ws["E1"] = "=ROUND(-PMT(B2/12/100,B3,B1),2)"

ws["D2"] = "Total Interest Paid"
ws["E2"] = "=SUM(C7:C606)"

ws["D3"] = "Total Extra Payments"
ws["E3"] = "=SUM(E7:E606)"

ws["D4"] = "Actual Term (Months)"
ws["E4"] = '=COUNTIF(H7:H606,">0")'

for cell in ["D1","D2","D3","D4"]:
    ws[cell].font = Font(bold=True)

# ======================
# TABLE HEADERS
# ======================

headers = [
    "Month No",
    "Opening Balance",
    "Interest",
    "Regular Payment",
    "Extra Payment",
    "Total Payment",
    "Principal Paid",
    "Closing Balance"
]

header_fill = PatternFill("solid", fgColor="1F4E78")

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=6, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = header_fill

# ======================
# AMORTIZATION TABLE
# ======================

MAX_ROWS = 600

for month in range(1, MAX_ROWS + 1):

    row = month + 6

    ws[f"A{row}"] = month

    if month == 1:
        ws[f"B{row}"] = "=B1"
    else:
        ws[f"B{row}"] = f"=H{row-1}"

    ws[f"C{row}"] = f"=IF(B{row}=0,0,B{row}*($B$2/12/100))"

    ws[f"D{row}"] = "=$E$1"

    ws[f"E{row}"] = 0

    ws[f"F{row}"] = f"=D{row}+E{row}"

    ws[f"G{row}"] = f"=F{row}-C{row}"

    ws[f"H{row}"] = f"=MAX(0,B{row}-G{row})"

# ======================
# FORMATTING
# ======================

for col in ws.columns:
    width = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws.column_dimensions[get_column_letter(col[0].column)].width = width + 5

currency_cols = ["B","C","D","E","F","G","H"]

for col in currency_cols:
    for row in range(7, MAX_ROWS + 7):
        ws[f"{col}{row}"].number_format = '#,##0.00'

# ======================
# CHART 1
# Outstanding Balance
# ======================

chart1 = LineChart()
chart1.title = "Outstanding Balance Trend"

data = Reference(ws,
                 min_col=8,
                 min_row=6,
                 max_row=126)

cats = Reference(ws,
                 min_col=1,
                 min_row=7,
                 max_row=126)

chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)

ws.add_chart(chart1, "J2")

# ======================
# CHART 2
# Interest vs Principal
# ======================

chart2 = LineChart()
chart2.title = "Principal vs Interest"

data2 = Reference(ws,
                  min_col=3,
                  max_col=7,
                  min_row=6,
                  max_row=126)

chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats)

ws.add_chart(chart2, "J20")

# ======================
# SAVE
# ======================

output_file = r"C:\Users\yarvi\Desktop\localsend\other\Professional_EMI_Calculator.xlsx"

wb.save(output_file)

print("SUCCESS")
print(output_file)
